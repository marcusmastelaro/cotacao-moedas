import requests
import schedule
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Função para consultar a cotação do dólar


def consultar_cotacao_dolar():
    try:
        url = "https://economia.awesomeapi.com.br/json/last/USD-BRL"
        resposta = requests.get(url)
        resposta.raise_for_status()
        dados = resposta.json()
        return dados['USDBRL']['bid']
    except requests.exceptions.RequestException as e:
        print(f"Erro ao consultar a cotação: {e}")
        return None

# Função para salvar os dados em uma planilha


def salvar_cotacao(cotacao):
    # Caminho da planilha
    caminho_arquivo = "C:\\Users\\marcus.mastelaro\\Documents\\Projetos-Python\\agentes\\cotacao_dolar\\cotacoes.xlsx"

    # Verificar se a planilha já existe
    if os.path.exists(caminho_arquivo):
        workbook = load_workbook(caminho_arquivo)
        sheet = workbook.active
    else:
        # Criar uma nova planilha
        workbook = Workbook()
        sheet = workbook.active
        # Criar cabeçalhos
        sheet.append(["Data", "Hora", "Cotação (R$)"])

    # Obter a data e a hora atuais
    agora = datetime.now()
    data = agora.strftime("%Y-%m-%d")
    hora = agora.strftime("%H:%M:%S")

    # Adicionar a cotação
    sheet.append([data, hora, cotacao])
    workbook.save(caminho_arquivo)
    print(f"Cotação salva: {data} {hora} - R${cotacao}")

# Função principal para buscar e salvar a cotação


def executar_agente():
    cotacao = consultar_cotacao_dolar()
    if cotacao:
        salvar_cotacao(cotacao)


# Agendar a execução a cada 10 minutos
schedule.every(5).minutes.do(executar_agente)

print("Agente iniciado. Consultando cotação a cada 5 minutos...")

# Loop para manter o agendamento em execução
while True:
    schedule.run_pending()
    time.sleep(1)
